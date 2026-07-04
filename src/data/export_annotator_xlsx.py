"""
Xuất data/vistqad/validation_sample_500.csv thành file .xlsx cho 2 người
thẩm định độc lập chấm điểm — 2 sheet giống hệt nhau (cùng thứ tự đã xáo
trộn), mỗi người chấm 1 sheet, KHÔNG thấy nhãn/ghi chú của người kia.

KHÔNG hiển thị auto_flag/auto_reasons trong sheet chấm (tránh neo/anchoring
bias — người thẩm định không nên biết trước máy đã nghi ngờ câu nào).
auto_flag chỉ dùng để đối chiếu SAU khi có kết quả chấm (không xuất ra đây).

3 cột chấm (1/0): dung_ngu_phap, dung_ngu_nghia, bao_toan_rang_buoc.
1 cột ghi_chu tự do.

Output: data/vistqad/validation_sample_500_for_annotators.xlsx
"""

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

DISPLAY_COLS = ["qid", "relation", "template_id", "question", "answer_label"]
GRADING_COLS = ["dung_ngu_phap", "dung_ngu_nghia", "bao_toan_rang_buoc", "ghi_chu"]


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    header = DISPLAY_COLS + GRADING_COLS
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for _, row in df.iterrows():
        ws.append([row[c] for c in DISPLAY_COLS] + ["", "", "", ""])

    # Data validation: 3 cột chấm chỉ nhận 0 hoặc 1
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv)
    n_rows = len(df) + 1
    for col_name in ["dung_ngu_phap", "dung_ngu_nghia", "bao_toan_rang_buoc"]:
        col_idx = header.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{n_rows}")

    widths = {"qid": 10, "relation": 12, "template_id": 22, "question": 60,
              "answer_label": 20, "dung_ngu_phap": 14, "dung_ngu_nghia": 15,
              "bao_toan_rang_buoc": 18, "ghi_chu": 30}
    for i, col_name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 15)
    ws.freeze_panes = "A2"


def main():
    src_path = os.path.join(ROOT, "data", "vistqad", "validation_sample_500.csv")
    df = pd.read_csv(src_path)

    # Xáo trộn thứ tự (cố định seed để tái lập), bỏ auto_flag/auto_reasons khỏi sheet chấm.
    shuffled = df.sample(frac=1, random_state=123).reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "annotator1", shuffled)
    _write_sheet(wb, "annotator2", shuffled)

    out_path = os.path.join(ROOT, "data", "vistqad", "validation_sample_500_for_annotators.xlsx")
    wb.save(out_path)
    print(f"Đã xuất {len(shuffled)} dòng x 2 sheet (annotator1, annotator2) -> {out_path}")
    print("LƯU Ý: 2 sheet cùng thứ tự câu hỏi (đã xáo trộn so với file gốc), "
          "KHÔNG hiển thị auto_flag — mỗi người chấm 1 sheet độc lập.")
    return out_path


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
