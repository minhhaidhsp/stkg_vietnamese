"""
Tính Cohen's κ từ data/vistqad/validation_sample_500_for_annotators.xlsx
SAU KHI 2 người đã chấm độc lập 2 sheet (annotator1, annotator2) — mỗi
sheet có 3 cột: dung_ngu_phap, dung_ngu_nghia, bao_toan_rang_buoc (1/0).

Tính κ RIÊNG cho từng tiêu chí + κ tổng hợp (coi 1 câu là "đạt" chỉ khi cả
3 tiêu chí đều =1, theo cả 2 người), cùng tỷ lệ đạt từng tiêu chí.

KHÔNG mô phỏng annotator — nếu sheet chưa được điền đầy đủ (còn ô rỗng),
script DỪNG và báo lỗi rõ ràng, không suy diễn số liệu.

Output: results/dataset/human_validation.json
"""

import json
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import cohen_kappa_score

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CRITERIA = ["dung_ngu_phap", "dung_ngu_nghia", "bao_toan_rang_buoc"]


def _load_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"LỖI: không tìm thấy sheet '{sheet_name}' trong {path}")
        sys.exit(1)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    return pd.DataFrame(data, columns=header)


def _check_complete(df: pd.DataFrame, sheet_name: str) -> None:
    for col in CRITERIA:
        n_missing = df[col].isna().sum()
        if n_missing > 0:
            print(f"CHƯA THỂ TÍNH κ: sheet '{sheet_name}', cột '{col}' còn "
                  f"{n_missing}/{len(df)} ô rỗng.")
            print("Cần người thẩm định điền đầy đủ (0/1) trước khi chạy lại.")
            sys.exit(1)


def main():
    path = os.path.join(ROOT, "data", "vistqad", "validation_sample_500_for_annotators.xlsx")
    if not os.path.exists(path):
        print(f"LỖI: chưa có file {path}. Chạy src/data/export_annotator_xlsx.py trước.")
        sys.exit(1)

    a1 = _load_sheet(path, "annotator1")
    a2 = _load_sheet(path, "annotator2")
    _check_complete(a1, "annotator1")
    _check_complete(a2, "annotator2")

    if len(a1) != len(a2) or not (a1["qid"].reset_index(drop=True) == a2["qid"].reset_index(drop=True)).all():
        print("LỖI: 2 sheet không cùng thứ tự/qid — không thể so khớp theo hàng.")
        sys.exit(1)

    report = {"n_samples": len(a1), "per_criterion": {}, "raw_agreement_per_criterion": {}}

    for col in CRITERIA:
        v1 = a1[col].astype(int)
        v2 = a2[col].astype(int)
        kappa = cohen_kappa_score(v1, v2)
        agreement = (v1 == v2).mean()
        pass_rate_1 = v1.mean()
        pass_rate_2 = v2.mean()
        report["per_criterion"][col] = {
            "cohen_kappa": round(float(kappa), 4),
            "raw_agreement_pct": round(float(agreement) * 100, 2),
            "pass_rate_annotator1_pct": round(float(pass_rate_1) * 100, 2),
            "pass_rate_annotator2_pct": round(float(pass_rate_2) * 100, 2),
        }

    # κ tổng hợp: "đạt" = cả 3 tiêu chí đều 1, riêng cho từng người, rồi so khớp
    combined_1 = (a1[CRITERIA].astype(int).sum(axis=1) == 3).astype(int)
    combined_2 = (a2[CRITERIA].astype(int).sum(axis=1) == 3).astype(int)
    kappa_combined = cohen_kappa_score(combined_1, combined_2)
    report["combined"] = {
        "definition": "đạt cả 3 tiêu chí (dung_ngu_phap=1 AND dung_ngu_nghia=1 AND bao_toan_rang_buoc=1)",
        "cohen_kappa": round(float(kappa_combined), 4),
        "raw_agreement_pct": round(float((combined_1 == combined_2).mean()) * 100, 2),
        "pass_rate_annotator1_pct": round(float(combined_1.mean()) * 100, 2),
        "pass_rate_annotator2_pct": round(float(combined_2.mean()) * 100, 2),
    }

    out_path = os.path.join(ROOT, "results", "dataset", "human_validation.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"\nĐã lưu -> {out_path}")
    return report


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
